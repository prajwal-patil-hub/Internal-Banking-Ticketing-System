"""Report generation service: CSV, Excel (.xlsx), PDF with comprehensive audit fields."""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone
from typing import Any

from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ValidationError
from app.core.logging import get_logger
from app.models.ticket import Ticket, TicketStatus
from app.models.user import User
from app.services.org_service import get_hierarchy_chain

log = get_logger(__name__)


REPORT_COLUMNS = [
    "ticket_id",
    "ticket_number",
    "title",
    "status",
    "priority",
    "source",
    "org_unit_code",
    "org_unit_name",
    "org_level",
    "hierarchy_chain",
    "raised_by_email",
    "raised_by_name",
    "raised_at",
    "acknowledged_at",
    "first_response_at",
    "resolved_by_email",
    "resolved_by_name",
    "resolved_at",
    "closed_at",
    "sla_breached",
    "sla_due_at",
    "sla_response_due_at",
    "escalation_count",
    "escalation_time_hrs",
    "reopen_count",
    "ai_assisted",
    "ai_confidence",
    "category",
    "subcategory",
    "tags",
    "department",
    "internal_notes",
]


async def _build_report_rows(
    db: AsyncSession,
    filters: dict,
    current_user: User,
) -> list[dict[str, Any]]:
    """Fetch tickets with all audit fields and return as list of dicts."""
    from app.models.escalation import EscalationEvent
    from app.models.ai_interaction import AIInteractionLog
    from app.services.org_service import get_accessible_org_unit_ids

    stmt = select(Ticket)

    # Apply org visibility
    if not current_user.is_super_admin:
        if current_user.org_unit_id:
            accessible = await get_accessible_org_unit_ids(current_user, db)
            if accessible is not None:
                from sqlalchemy import or_
                stmt = stmt.where(
                    or_(
                        Ticket.org_unit_id.in_([str(uid) for uid in accessible]),
                        Ticket.assignee_id == current_user.id,
                    )
                )
        elif current_user.role.name == "branch_user":
            stmt = stmt.where(Ticket.reporter_id == current_user.id)

    # Apply date filters
    if filters.get("from_date"):
        stmt = stmt.where(Ticket.created_at >= filters["from_date"])
    if filters.get("to_date"):
        stmt = stmt.where(Ticket.created_at <= filters["to_date"])
    if filters.get("status"):
        try:
            stmt = stmt.where(Ticket.status == TicketStatus(filters["status"]))
        except ValueError:
            pass
    if filters.get("org_unit_id"):
        stmt = stmt.where(Ticket.org_unit_id == uuid.UUID(str(filters["org_unit_id"])))
    if filters.get("priority"):
        from app.models.ticket import TicketPriority
        try:
            stmt = stmt.where(Ticket.priority == TicketPriority(filters["priority"]))
        except ValueError:
            pass

    stmt = stmt.order_by(Ticket.created_at.desc())
    result = await db.execute(stmt)
    tickets = result.scalars().all()

    rows: list[dict[str, Any]] = []
    for ticket in tickets:
        # Escalation count and total time
        esc_result = await db.execute(
            select(func.count(EscalationEvent.id)).where(
                EscalationEvent.ticket_id == ticket.id
            )
        )
        esc_count = esc_result.scalar_one() or 0

        # Calculate escalation time (sum of durations between escalations)
        esc_time_hrs: float | None = None
        if esc_count > 0:
            esc_events_result = await db.execute(
                select(EscalationEvent.triggered_at, EscalationEvent.resolved_at).where(
                    EscalationEvent.ticket_id == ticket.id
                ).order_by(EscalationEvent.triggered_at)
            )
            esc_events = esc_events_result.all()
            total_esc_ms = 0.0
            for ev in esc_events:
                if ev.triggered_at and ev.resolved_at:
                    total_esc_ms += (ev.resolved_at - ev.triggered_at).total_seconds()
            esc_time_hrs = round(total_esc_ms / 3600, 2) if total_esc_ms else None

        # AI assisted check
        ai_result = await db.execute(
            select(func.count(AIInteractionLog.id)).where(
                AIInteractionLog.ticket_id == ticket.id
            )
        )
        ai_assisted = (ai_result.scalar_one() or 0) > 0 or bool(ticket.ai_category)

        # Hierarchy chain
        hierarchy_chain_str = ""
        if ticket.org_unit_id:
            chain = await get_hierarchy_chain(db, ticket.org_unit_id)
            hierarchy_chain_str = " > ".join(
                f"{c['name']} ({c['code']})" for c in chain
            )

        rows.append({
            "ticket_id": str(ticket.id),
            "ticket_number": ticket.ticket_number,
            "title": ticket.title,
            "status": ticket.status.value,
            "priority": ticket.priority.value,
            "source": ticket.source.value,
            "org_unit_code": ticket.org_unit.code if ticket.org_unit else "",
            "org_unit_name": ticket.org_unit.name if ticket.org_unit else "",
            "org_level": (
                ticket.org_unit.hierarchy_level.name
                if ticket.org_unit and ticket.org_unit.hierarchy_level
                else ""
            ),
            "hierarchy_chain": hierarchy_chain_str,
            "raised_by_email": ticket.reporter.email if ticket.reporter else "",
            "raised_by_name": ticket.reporter.full_name if ticket.reporter else "",
            "raised_at": ticket.created_at.isoformat() if ticket.created_at else "",
            "acknowledged_at": "",
            "first_response_at": ticket.first_response_at.isoformat() if ticket.first_response_at else "",
            "resolved_by_email": ticket.assignee.email if ticket.assignee else "",
            "resolved_by_name": ticket.assignee.full_name if ticket.assignee else "",
            "resolved_at": ticket.resolved_at.isoformat() if ticket.resolved_at else "",
            "closed_at": ticket.closed_at.isoformat() if ticket.closed_at else "",
            "sla_breached": "Yes" if ticket.sla_breached else "No",
            "sla_due_at": ticket.resolution_due_at.isoformat() if ticket.resolution_due_at else "",
            "sla_response_due_at": ticket.response_due_at.isoformat() if ticket.response_due_at else "",
            "escalation_count": esc_count,
            "escalation_time_hrs": esc_time_hrs if esc_time_hrs is not None else "",
            "reopen_count": ticket.reopen_count or 0,
            "ai_assisted": "Yes" if ai_assisted else "No",
            "ai_confidence": round(ticket.ai_confidence * 100, 1) if ticket.ai_confidence else "",
            "category": ticket.category.name if ticket.category else "",
            "subcategory": ticket.subcategory.name if ticket.subcategory else "",
            "tags": ", ".join(ticket.tags or []),
            "department": ticket.department or "",
            "internal_notes": ticket.internal_notes or "",
        })

    return rows


async def generate_csv(db: AsyncSession, filters: dict, current_user: User) -> bytes:
    rows = await _build_report_rows(db, filters, current_user)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=REPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


async def generate_excel(db: AsyncSession, filters: dict, current_user: User) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = await _build_report_rows(db, filters, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ticket Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A5276")

    headers = [col.replace("_", " ").title() for col in REPORT_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_key in enumerate(REPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(col_key, "")))

    # Auto-width (approximate)
    for col_idx in range(1, len(REPORT_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_pdf(db: AsyncSession, filters: dict, current_user: User) -> bytes:
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    rows = await _build_report_rows(db, filters, current_user)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A3),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    story = [
        Paragraph("Ticket Audit Report", styles["Title"]),
        Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | "
            f"Total records: {len(rows)}",
            styles["Normal"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    # Only show key columns in PDF (full list is too wide for a page)
    pdf_cols = [
        "ticket_number", "title", "status", "priority",
        "org_unit_code", "raised_by_name", "raised_at",
        "resolved_at", "sla_breached", "reopen_count",
        "escalation_count", "ai_assisted",
    ]
    headers = [[col.replace("_", " ").title() for col in pdf_cols]]
    data_rows = [
        [str(row.get(col, ""))[:40] for col in pdf_cols]
        for row in rows
    ]
    table_data = headers + data_rows

    col_widths = [3.5 * cm] * len(pdf_cols)
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#AED6F1")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Analytics export — charts and KPI tiles
# ---------------------------------------------------------------------------
#
# The dashboard could only save a chart as a PNG, produced in the browser from
# the SVG. These build the other two formats server-side using the libraries
# the ticket report already depends on, so the frontend needs no new packages:
# a PDF that embeds the rendered chart image above its data, and a workbook
# with one sheet per chart plus the KPI tiles.


def _decode_chart_png(image_data_url: str | None) -> bytes | None:
    """Pull raw PNG bytes out of a `data:image/png;base64,...` URL."""
    if not image_data_url:
        return None
    import base64

    _, _, encoded = image_data_url.partition("base64,")
    if not encoded:
        return None
    try:
        # validate=True so junk raises instead of silently decoding to b'' —
        # empty bytes would sail past this guard and fail deeper in the PDF
        # builder, where the cause is much harder to see.
        decoded = base64.b64decode(encoded, validate=True)
    except Exception:  # noqa: BLE001 - a bad image must not fail the export
        log.warning("chart_export.bad_image_payload")
        return None
    return decoded or None


def _chart_grid(chart: Any) -> tuple[list[str], list[list[Any]]]:
    """Normalise one chart's rows into (columns, rows-as-lists).

    Rows arrive keyed by column — that is what Recharts holds and what the
    frontend sends. A list-of-lists is accepted too, because it is the obvious
    shape for anyone calling this endpoint by hand and there is no reason to
    reject it.

    Anything genuinely unusable raises ValidationError rather than propagating
    an AttributeError: the route takes a free-form dict, so a malformed payload
    is a client error and must not read as a server fault.
    """
    if not isinstance(chart, dict):
        raise ValidationError("Each chart must be an object.")

    rows = chart.get("rows") or []
    if not isinstance(rows, list):
        raise ValidationError(
            f"Chart '{chart.get('title') or 'untitled'}': rows must be a list."
        )
    if not rows:
        return [], []

    declared = chart.get("columns")
    columns = [str(c) for c in declared] if declared else []

    first = rows[0]
    if isinstance(first, dict):
        if not columns:
            columns = [str(k) for k in first]
        grid = [
            [row.get(c, "") for c in columns]
            for row in rows
            if isinstance(row, dict)
        ]
        return columns, grid

    if isinstance(first, (list, tuple)):
        if not columns:
            columns = [f"Column {i + 1}" for i in range(len(first))]
        grid = [
            # Pad or trim so a ragged row cannot misalign the whole table.
            [*row, *([""] * (len(columns) - len(row)))][: len(columns)]
            for row in rows
            if isinstance(row, (list, tuple))
        ]
        return columns, grid

    raise ValidationError(
        f"Chart '{chart.get('title') or 'untitled'}': each row must be an "
        f"object or a list, not {type(first).__name__}."
    )


def generate_analytics_excel(payload: dict) -> bytes:
    """Workbook: a KPI summary sheet, then one sheet per chart's data."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A5276")
    title_font = Font(bold=True, size=13)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; every sheet below is named

    kpis = payload.get("kpis") or []
    if kpis:
        ws = wb.create_sheet("Summary")
        ws["A1"] = payload.get("title") or "Dashboard Summary"
        ws["A1"].font = title_font
        if generated := payload.get("generated_at"):
            ws["A2"] = f"Generated {generated}"
        for col, header in enumerate(("Metric", "Value"), start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, kpi in enumerate(kpis, start=5):
            ws.cell(row=row_idx, column=1, value=str(kpi.get("label", "")))
            ws.cell(row=row_idx, column=2, value=kpi.get("value"))
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 18

    for index, chart in enumerate(payload.get("charts") or [], start=1):
        # Excel sheet names cap at 31 chars and reject : \ / ? * [ ]
        raw_name = str(chart.get("title") or f"Chart {index}")
        safe = "".join("-" if ch in ':\\/?*[]' else ch for ch in raw_name)[:31] or f"Chart {index}"
        while safe in wb.sheetnames:  # titles need not be unique; sheets do
            safe = f"{safe[:28]}_{index}"
        ws = wb.create_sheet(safe)

        columns, rows = _chart_grid(chart)

        ws["A1"] = raw_name
        ws["A1"].font = title_font
        for col_idx, key in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=str(key).replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, start=4):
            for col_idx, value in enumerate(row, start=1):
                # openpyxl only accepts scalars; anything nested becomes text.
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value if isinstance(value, (str, int, float, bool, type(None)))
                    else str(value),
                )
        for col_idx in range(1, max(len(columns), 1) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        if rows:
            ws.freeze_panes = "A4"

    if not wb.sheetnames:  # nothing supplied — still return a valid workbook
        wb.create_sheet("Empty")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_analytics_pdf(payload: dict) -> bytes:
    """PDF: each chart's rendered image above the numbers behind it."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (
        Image,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        topMargin=1.4 * cm, bottomMargin=1.4 * cm,
    )
    usable_width = doc.width
    story: list = []

    story.append(Paragraph(payload.get("title") or "Dashboard Report", styles["Title"]))
    if generated := payload.get("generated_at"):
        story.append(Paragraph(f"Generated {generated}", styles["Normal"]))
    story.append(Spacer(1, 0.6 * cm))

    if kpis := payload.get("kpis") or []:
        story.append(Paragraph("Key metrics", styles["Heading2"]))
        table = Table(
            [["Metric", "Value"]] + [[str(k.get("label", "")), str(k.get("value", ""))] for k in kpis],
            colWidths=[usable_width * 0.62, usable_width * 0.38],
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0B7C3")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.7 * cm))

    charts = payload.get("charts") or []
    for index, chart in enumerate(charts):
        story.append(Paragraph(str(chart.get("title") or f"Chart {index + 1}"), styles["Heading2"]))
        story.append(Spacer(1, 0.25 * cm))

        if image_bytes := _decode_chart_png(chart.get("image")):
            try:
                reader = ImageReader(io.BytesIO(image_bytes))
                src_w, src_h = reader.getSize()
                # Scale to the text column, capping height so the data table
                # below still shares the page.
                width = min(usable_width, src_w)
                height = width * src_h / src_w
                max_height = 9 * cm
                if height > max_height:
                    height, width = max_height, max_height * src_w / src_h
                story.append(Image(io.BytesIO(image_bytes), width=width, height=height))
                story.append(Spacer(1, 0.4 * cm))
            except Exception:  # noqa: BLE001 - fall back to the table alone
                log.warning("chart_export.image_render_failed", chart=chart.get("title"))

        columns, rows = _chart_grid(chart)
        if rows and columns:
            header = [str(c).replace("_", " ").title() for c in columns]
            body = [[str(v) for v in row] for row in rows]
            table = Table([header] + body, colWidths=[usable_width / len(columns)] * len(columns))
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0B7C3")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F8")]),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(table)

        if index < len(charts) - 1:
            story.append(PageBreak())

    if not story:
        story.append(Paragraph("No data available.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()
