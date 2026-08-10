"""Dashboard analytics export — PDF and Excel generation."""

from __future__ import annotations

import base64
import io
import struct
import zlib

import pytest
from openpyxl import load_workbook

from app.services.report_service import (
    _decode_chart_png,
    generate_analytics_excel,
    generate_analytics_pdf,
)


def _png_bytes(width: int = 4, height: int = 4) -> bytes:
    """A real, minimal PNG so the image path is genuinely exercised."""
    def chunk(tag: bytes, data: bytes) -> bytes:
        body = tag + data
        return struct.pack(">I", len(data)) + body + struct.pack(">I", zlib.crc32(body) & 0xFFFFFFFF)

    raw = b"".join(b"\x00" + b"\xff\x00\x00" * width for _ in range(height))
    return (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0))
        + chunk(b"IDAT", zlib.compress(raw))
        + chunk(b"IEND", b"")
    )


def _data_url() -> str:
    return "data:image/png;base64," + base64.b64encode(_png_bytes()).decode()


PAYLOAD = {
    "title": "Ticket Analytics",
    "generated_at": "2026-08-10 12:00 UTC",
    "kpis": [
        {"label": "Total tickets", "value": 21},
        {"label": "SLA breached", "value": 7},
    ],
    "charts": [
        {
            "title": "Tickets by Status",
            "columns": ["name", "count"],
            "rows": [{"name": "new", "count": 3}, {"name": "escalated", "count": 3}],
            "image": _data_url(),
        },
        {
            "title": "Tickets by Priority",
            "columns": ["name", "value"],
            "rows": [{"name": "critical", "value": 3}],
        },
    ],
}


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def test_pdf_is_a_real_pdf() -> None:
    out = generate_analytics_pdf(PAYLOAD)

    assert out.startswith(b"%PDF-")
    assert out.rstrip().endswith(b"%%EOF")


def test_pdf_survives_a_corrupt_chart_image() -> None:
    """A bad image must cost the picture, not the whole document."""
    payload = {"charts": [{"title": "X", "rows": [{"a": 1}], "image": "data:image/png;base64,!!nope!!"}]}

    out = generate_analytics_pdf(payload)

    assert out.startswith(b"%PDF-")


def test_pdf_handles_an_empty_payload() -> None:
    out = generate_analytics_pdf({})

    assert out.startswith(b"%PDF-")


def test_pdf_handles_charts_with_no_rows() -> None:
    out = generate_analytics_pdf({"charts": [{"title": "Empty", "rows": []}]})

    assert out.startswith(b"%PDF-")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def test_excel_has_a_summary_sheet_and_one_sheet_per_chart() -> None:
    wb = load_workbook(io.BytesIO(generate_analytics_excel(PAYLOAD)))

    assert wb.sheetnames == ["Summary", "Tickets by Status", "Tickets by Priority"]


def test_excel_summary_carries_the_kpi_values() -> None:
    wb = load_workbook(io.BytesIO(generate_analytics_excel(PAYLOAD)))
    ws = wb["Summary"]

    assert (ws.cell(5, 1).value, ws.cell(5, 2).value) == ("Total tickets", 21)
    assert (ws.cell(6, 1).value, ws.cell(6, 2).value) == ("SLA breached", 7)


def test_excel_chart_sheet_carries_headers_and_rows() -> None:
    wb = load_workbook(io.BytesIO(generate_analytics_excel(PAYLOAD)))
    ws = wb["Tickets by Status"]

    assert (ws.cell(3, 1).value, ws.cell(3, 2).value) == ("Name", "Count")
    assert (ws.cell(4, 1).value, ws.cell(4, 2).value) == ("new", 3)
    assert (ws.cell(5, 1).value, ws.cell(5, 2).value) == ("escalated", 3)


def test_excel_deduplicates_repeated_chart_titles() -> None:
    """Chart titles need not be unique; worksheet names must be."""
    payload = {"charts": [
        {"title": "Same", "rows": [{"a": 1}]},
        {"title": "Same", "rows": [{"a": 2}]},
    ]}

    wb = load_workbook(io.BytesIO(generate_analytics_excel(payload)))

    assert len(wb.sheetnames) == 2
    assert len(set(wb.sheetnames)) == 2


def test_excel_sanitises_illegal_sheet_names() -> None:
    """Excel rejects : \\ / ? * [ ] and caps names at 31 characters."""
    payload = {"charts": [{"title": "A" * 50 + "/bad:name?", "rows": [{"a": 1}]}]}

    wb = load_workbook(io.BytesIO(generate_analytics_excel(payload)))

    name = wb.sheetnames[0]
    assert len(name) <= 31
    assert not set(name) & set(':\\/?*[]')


def test_excel_infers_columns_when_not_given() -> None:
    payload = {"charts": [{"title": "Inferred", "rows": [{"alpha": 1, "beta": 2}]}]}

    ws = load_workbook(io.BytesIO(generate_analytics_excel(payload)))["Inferred"]

    assert {ws.cell(3, 1).value, ws.cell(3, 2).value} == {"Alpha", "Beta"}


def test_excel_is_valid_with_an_empty_payload() -> None:
    """openpyxl cannot save a workbook with no sheets at all."""
    wb = load_workbook(io.BytesIO(generate_analytics_excel({})))

    assert wb.sheetnames


# ---------------------------------------------------------------------------
# Image decoding
# ---------------------------------------------------------------------------

def test_decode_chart_png_round_trips() -> None:
    assert _decode_chart_png(_data_url()) == _png_bytes()


@pytest.mark.parametrize("value", [None, "", "notadataurl", "data:image/png;base64,%%%"])
def test_decode_chart_png_returns_none_on_bad_input(value) -> None:
    assert _decode_chart_png(value) is None
